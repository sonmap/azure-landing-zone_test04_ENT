#!/usr/bin/env python3
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def style_sheet(ws, color: str) -> None:
    thin = Side(style="thin", color="D9D9D9")
    fill = PatternFill("solid", fgColor=color)
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(
            62, max(14, len(str(ws.cell(1, col).value or "")) + 5)
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_sheet(wb, name: str, color: str, headers, rows) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws, color)


def append_mapping(wb, sheet_names) -> None:
    if "19_TFVars_Mapping" not in wb.sheetnames:
        return
    mapping = wb["19_TFVars_Mapping"]
    existing = {str(mapping.cell(r, 4).value) for r in range(2, mapping.max_row + 1)}
    number = mapping.max_row
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for row_idx in range(2, ws.max_row + 1):
            row = {headers[c - 1]: ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)}
            tfvars_path = row.get("tfvars path")
            command = row.get("선택 갱신 명령")
            if not tfvars_path or str(tfvars_path) in existing:
                continue
            condition = []
            for key in [
                "부서",
                "업무",
                "Environment",
                "Request Name",
                "Pipeline Name",
                "Guard Name",
                "Scope",
                "Target",
                "Data Domain",
                "Control Name",
                "Test Name",
                "Account Name",
            ]:
                if row.get(key):
                    condition.append(f"{key}={row[key]}")
            number += 1
            mapping.append([
                number,
                sheet_name,
                ", ".join(condition),
                tfvars_path,
                command or "",
                str(tfvars_path).rsplit("/terraform.tfvars", 1)[0],
                row.get("비고", ""),
            ])


def main() -> None:
    path = Path("azure_landingzone_terraform_module_design.xlsx")
    wb = openpyxl.load_workbook(path)

    specs = {
        "34_LZ_Vending_Request": (
            "1F4E78",
            [
                "ModuleID", "Request Name", "부서", "업무", "Environment", "Subscription Alias",
                "Management Group", "ResourceGroup", "Region", "VNet", "CIDR", "Policy Baseline",
                "RBAC Group", "Budget", "CostCenter", "Owner", "Connectivity", "Private DNS",
                "tfvars path", "선택 갱신 명령", "Status", "비고",
            ],
            [
                ["MOD-090", "lz-sales-dev", "Sales", "영업지원", "dev", "sub-sl-sales-dev", "LandingZones/Corp/Dev", "rg-sl-sales-dev", "koreacentral", "vnet-sl-sales-dev", "10.40.0.0/20", "SamsungLife-Secure-LZ", "grp-az-sales-dev-owner", "1시간 검증용", "CC10010", "sales-it", "Hub Peering", "Central", "live/00-foundation/lz-vending-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target lz-vending --department Sales --environment dev", "Draft", "신규 Landing Zone 자동 발급 요청"],
                ["MOD-090", "lz-ai-sandbox", "AI Platform", "AI Sandbox", "sandbox", "sub-sl-ai-sandbox", "LandingZones/Sandbox", "rg-sl-ai-sandbox", "koreacentral", "vnet-sl-ai-sandbox", "10.41.0.0/20", "SamsungLife-AI-Sandbox", "grp-az-ai-sandbox-owner", "월예산 입력", "CC90010", "ai-platform", "Hub Peering", "Central", "live/00-foundation/lz-vending-ai-sandbox/terraform.tfvars", "tools/update_extended_tfvars.sh --target lz-vending --department \"AI Platform\" --environment sandbox", "Draft", "AI Sandbox Landing Zone 발급"],
            ],
        ),
        "35_Pipeline_Governance": (
            "548235",
            [
                "ModuleID", "Pipeline Name", "대상", "Trigger", "Plan Required", "Policy Check",
                "Cost Check", "Security Check", "Approver", "Apply Window", "TTL Required",
                "Artifact Retention", "State Backend", "tfvars path", "선택 갱신 명령", "Status", "비고",
            ],
            [
                ["MOD-091", "pipe-lz-platform", "Platform Landing Zone", "PR/Merge", "Y", "Checkov/Azure Policy", "Infracost/Budget", "tfsec/Defender", "platform-owner", "CAB 승인 후", "Y", "180d", "azurerm", "live/00-foundation/pipeline-governance/terraform.tfvars", "tools/update_extended_tfvars.sh --target pipeline-governance", "Draft", "Plan/Policy/Cost/Security gate"],
                ["MOD-091", "pipe-workload-sales-dev", "Sales dev workload", "PR/Merge", "Y", "Azure Policy", "Budget threshold", "Secret scan/Image scan", "sales-owner", "09-18", "Y", "90d", "azurerm", "live/30-services/pipeline-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target pipeline-governance --department Sales --environment dev", "Draft", "업무별 배포 게이트"],
            ],
        ),
        "36_FinOps_CostGuard": (
            "9E480E",
            [
                "ModuleID", "Guard Name", "부서", "업무", "Environment", "Scope", "Monthly Budget",
                "Alert 50%", "Alert 80%", "Alert 100%", "Allowed SKU", "Denied SKU", "High Cost Approval",
                "Auto TTL", "Owner", "tfvars path", "선택 갱신 명령", "Status", "비고",
            ],
            [
                ["MOD-092", "cost-sales-dev", "Sales", "영업지원", "dev", "rg-sl-sales-dev", "100000 KRW", "Y", "Y", "Y", "D2s_v3,D4s_v5,Basic/S0", "M-series,AppGw WAF_v2 without approval", "Y", "3600s", "sales-it", "live/30-services/finops-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target finops --department Sales --environment dev", "Draft", "1시간 검증 비용 보호"],
                ["MOD-092", "cost-platform-prod", "Platform", "공통", "prod", "subscription", "월예산 입력", "Y", "Y", "Y", "Approved SKU only", "Unapproved public/expensive SKU", "Y", "N", "platform-owner", "live/10-platform/finops-platform/terraform.tfvars", "tools/update_extended_tfvars.sh --target finops --department Platform --environment prod", "Draft", "프로덕션 예산/sku 통제"],
            ],
        ),
        "37_Defender_CSPM": (
            "C00000",
            [
                "ModuleID", "Scope", "Environment", "Defender Plan", "CSPM", "Regulatory Standard",
                "Secure Score Target", "Auto Provisioning", "JIT", "Vulnerability Assessment",
                "DevOps Security", "Owner", "tfvars path", "선택 갱신 명령", "Status", "비고",
            ],
            [
                ["MOD-093", "subscription", "prod", "Servers/Containers/Storage/KeyVault", "Foundational + Defender CSPM", "MCSB", "80%", "Y", "Y", "Y", "Y", "security-owner", "live/10-platform/defender-cspm/terraform.tfvars", "tools/update_extended_tfvars.sh --target defender-cspm --environment prod", "Draft", "Microsoft Cloud Security Benchmark 기준"],
                ["MOD-093", "rg-sl-sales-dev", "dev", "Servers/Containers", "Foundational CSPM", "MCSB", "70%", "Y", "N", "Y", "Y", "sales-security", "live/30-services/defender-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target defender-cspm --department Sales --environment dev", "Draft", "개발/검증 보안 태세"],
            ],
        ),
        "38_AzureArc_Hybrid": (
            "31859B",
            [
                "ModuleID", "Target", "부서", "업무", "Environment", "On-prem Location", "Arc Type",
                "ResourceGroup", "Connectivity", "Private Link Scope", "Update Manager", "Defender",
                "Identity", "tfvars path", "선택 갱신 명령", "Status", "비고",
            ],
            [
                ["MOD-094", "onprem-sales-web", "Sales", "영업지원", "dev", "IDC-Seoul", "Arc-enabled servers", "rg-sl-sales-dev", "ExpressRoute", "ampls-sl-platform", "Y", "Y", "SystemAssigned", "live/30-services/arc-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target azure-arc --workload 영업지원 --environment dev", "Draft", "온프레미스 서버 통합 관리"],
                ["MOD-094", "onprem-aks-ai", "AI Platform", "AI Sandbox", "sandbox", "IDC-Seoul", "Arc-enabled Kubernetes", "rg-sl-ai-sandbox", "ExpressRoute", "ampls-sl-platform", "Y", "Y", "Workload Identity", "live/30-services/arc-ai-sandbox/terraform.tfvars", "tools/update_extended_tfvars.sh --target azure-arc --workload \"AI Sandbox\" --environment sandbox", "Draft", "온프레미스 K8s 관리"],
            ],
        ),
        "39_Data_Governance": (
            "7030A0",
            [
                "ModuleID", "Data Domain", "부서", "업무", "Environment", "Classification", "Purview",
                "DLP", "Retention", "Encryption", "Data Owner", "Access Approval", "Lineage",
                "tfvars path", "선택 갱신 명령", "Status", "비고",
            ],
            [
                ["MOD-095", "sales-customer", "Sales", "영업지원", "dev", "restricted", "Y", "Y", "3y", "CMK", "sales-data-owner", "Y", "Y", "live/30-services/data-gov-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target data-governance --workload 영업지원 --environment dev", "Draft", "고객/영업 데이터 통제"],
                ["MOD-095", "ai-rag-docs", "AI Platform", "AI Sandbox", "sandbox", "internal/restricted", "Y", "Y", "1y", "CMK", "ai-data-owner", "Y", "Y", "live/30-services/data-gov-ai-sandbox/terraform.tfvars", "tools/update_extended_tfvars.sh --target data-governance --workload \"AI Sandbox\" --environment sandbox", "Draft", "RAG 데이터 분류/계보"],
            ],
        ),
        "40_DevSecOps_SupplyChain": (
            "00B050",
            [
                "ModuleID", "Control Name", "부서", "업무", "Environment", "Repo/Registry", "Image Scan",
                "SBOM", "Signing", "Secret Scan", "Dependency Scan", "Policy Gate", "Block Critical",
                "tfvars path", "선택 갱신 명령", "Status", "비고",
            ],
            [
                ["MOD-096", "supplychain-sales-api", "Sales", "영업지원", "dev", "acrslplatform/sales-api", "Y", "Y", "Y", "Y", "Y", "Y", "Y", "live/30-services/devsecops-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target devsecops --workload 영업지원 --environment dev", "Draft", "Container Apps/AKS 이미지 보안"],
                ["MOD-096", "supplychain-ai-rag", "AI Platform", "AI Sandbox", "sandbox", "acrslplatform/rag-api", "Y", "Y", "Y", "Y", "Y", "Y", "Y", "live/30-services/devsecops-ai-sandbox/terraform.tfvars", "tools/update_extended_tfvars.sh --target devsecops --workload \"AI Sandbox\" --environment sandbox", "Draft", "AI API 공급망 보안"],
            ],
        ),
        "41_Resilience_DRTest": (
            "5F497A",
            [
                "ModuleID", "Test Name", "부서", "업무", "Environment", "Scenario", "RTO Target",
                "RPO Target", "Frequency", "Failover Region", "Runbook", "Owner", "Evidence",
                "tfvars path", "선택 갱신 명령", "Status", "비고",
            ],
            [
                ["MOD-097", "drtest-sales-dev", "Sales", "영업지원", "dev", "VM restore + LB health", "4h", "24h", "Quarterly", "koreasouth", "RUN-DR-SALES-001", "sales-owner", "Required", "live/30-services/drtest-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target dr-test --workload 영업지원 --environment dev", "Draft", "복구훈련/증적"],
                ["MOD-097", "drtest-ai-sandbox", "AI Platform", "AI Sandbox", "sandbox", "Storage restore + private endpoint validation", "8h", "24h", "Semiannual", "koreasouth", "RUN-DR-AI-001", "ai-owner", "Required", "live/30-services/drtest-ai-sandbox/terraform.tfvars", "tools/update_extended_tfvars.sh --target dr-test --workload \"AI Sandbox\" --environment sandbox", "Draft", "AI 데이터 복구 검증"],
            ],
        ),
        "42_Emergency_Access": (
            "595959",
            [
                "ModuleID", "Account Name", "Purpose", "Scope", "MFA", "Conditional Access Exception",
                "Credential Storage", "Rotation Days", "Use Approval", "Post Review", "Alert",
                "Owner", "tfvars path", "선택 갱신 명령", "Status", "비고",
            ],
            [
                ["MOD-098", "breakglass-az-001", "Tenant emergency access", "Tenant Root/Platform", "Y", "Minimal documented exception", "HSM/Offline vault", 90, "Y", "Y", "Y", "security-owner", "live/00-foundation/emergency-access/terraform.tfvars", "tools/update_extended_tfvars.sh --target emergency-access", "Draft", "비상 접근 계정"],
                ["MOD-098", "breakglass-ops-001", "Ops emergency access", "Platform subscriptions", "Y", "Minimal documented exception", "HSM/Offline vault", 90, "Y", "Y", "Y", "platform-owner", "live/00-foundation/emergency-access/terraform.tfvars", "tools/update_extended_tfvars.sh --target emergency-access", "Draft", "운영 비상 계정"],
            ],
        ),
    }

    for name, (color, headers, rows) in specs.items():
        add_sheet(wb, name, color, headers, rows)

    append_mapping(wb, specs.keys())
    wb.save(path)
    print("latest landing zone trend sheets added")


if __name__ == "__main__":
    main()
