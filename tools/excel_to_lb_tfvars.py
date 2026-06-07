#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List

import openpyxl


def rows_as_dicts(ws) -> List[Dict[str, Any]]:
    headers = None
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if not any(v is not None for v in values):
            continue
        if headers is None:
            normalized = {str(v).strip().lower() for v in values if v is not None}
            if {"moduleid", "lb type", "name"}.issubset(normalized):
                headers = [str(v).strip() if v is not None else f"col_{i}" for i, v in enumerate(values)]
            continue
        item = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i] and values[i] is not None}
        if item:
            rows.append(item)
    return rows


def value(row: Dict[str, Any], key: str, default: Any = None) -> Any:
    for k, v in row.items():
        if str(k).strip().lower() == key.strip().lower() and v not in (None, ""):
            return v
    return default


def slug(raw: Any) -> str:
    text = str(raw or "").strip().lower()
    replacements = {
        "영업지원": "sales",
        "영업": "sales",
        "계약관리": "policy",
        "모바일api": "digital",
        "상품": "product",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "item"


def hcl_key(key: str) -> str:
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", key):
        return key
    return json.dumps(key, ensure_ascii=False)


def to_hcl(obj: Any, indent: int = 0) -> str:
    pad = " " * indent
    child = " " * (indent + 2)
    if isinstance(obj, bool):
        return "true" if obj else "false"
    if isinstance(obj, (int, float)):
        return str(obj)
    if isinstance(obj, str):
        return json.dumps(obj, ensure_ascii=False)
    if isinstance(obj, list):
        if not obj:
            return "[]"
        return "[\n" + ",\n".join(f"{child}{to_hcl(v, indent + 2)}" for v in obj) + f"\n{pad}]"
    if isinstance(obj, dict):
        if not obj:
            return "{}"
        lines = ["{"]
        for key in sorted(obj):
            lines.append(f"{child}{hcl_key(str(key))} = {to_hcl(obj[key], indent + 2)}")
        lines.append(f"{pad}}}")
        return "\n".join(lines)
    return json.dumps(str(obj), ensure_ascii=False)


def write_tfvars(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n\n".join(f"{k} = {to_hcl(data[k])}" for k in sorted(data)) + "\n", encoding="utf-8")
    print(f"created: {path}")


def selected(row: Dict[str, Any], args: argparse.Namespace) -> bool:
    if args.target and str(value(row, "LB Type", "")).strip().lower() != args.target.lower():
        return False
    if args.department and slug(value(row, "부서", "")) != slug(args.department):
        return False
    if args.environment and str(value(row, "Environment", "")).strip().lower() != args.environment.lower():
        return False
    if args.workload and slug(value(row, "업무", "")) != slug(args.workload):
        return False
    return True


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="azure_landingzone_terraform_module_design.xlsx")
    parser.add_argument("--out", default="live")
    parser.add_argument("--target", choices=["alb", "nlb"])
    parser.add_argument("--department")
    parser.add_argument("--environment")
    parser.add_argument("--workload")
    parser.add_argument("--location", default="koreacentral")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    if "20_LoadBalancer" not in wb.sheetnames:
        raise SystemExit("20_LoadBalancer 시트가 없습니다.")

    for row in rows_as_dicts(wb["20_LoadBalancer"]):
        if not selected(row, args):
            continue
        lb_type = str(value(row, "LB Type")).strip().lower()
        workload = slug(value(row, "업무"))
        env = str(value(row, "Environment")).strip().lower()
        name = str(value(row, "Name"))
        rg = str(value(row, "ResourceGroup"))
        vnet = str(value(row, "VNet"))
        subnet = str(value(row, "Subnet"))
        frontend_ip = str(value(row, "Frontend Private IP"))
        backend = str(value(row, "Backend Target"))
        probe_protocol = str(value(row, "Probe Protocol", "Tcp"))
        probe_port = int(value(row, "Probe Port", 80))
        rule_protocol = str(value(row, "Rule Protocol", "Tcp"))
        frontend_port = int(value(row, "Frontend Port", 80))
        backend_port = int(value(row, "Backend Port", 80))
        tags = {
            "project": "samsunglife-landingzone",
            "managed_by": "terraform",
            "lb_type": lb_type,
            "environment": env,
        }

        if lb_type == "nlb":
            write_tfvars(
                Path(args.out) / "30-services" / f"nlb-{workload}-{env}-web" / "terraform.tfvars",
                {
                    "resource_group_name": rg,
                    "location": args.location,
                    "vnet_resource_group_name": rg,
                    "vnet_name": vnet,
                    "subnet_name": subnet,
                    "lb_name": name,
                    "frontend_ip_name": "fe-web",
                    "frontend_private_ip": frontend_ip,
                    "backend_addresses": {"web01": {"name": "web01", "ip_address": "10.40.1.10" if backend.startswith("vm-") else backend}},
                    "probe": {"name": "probe-web", "protocol": probe_protocol, "port": probe_port},
                    "rule": {"name": "rule-web", "protocol": rule_protocol, "frontend_port": frontend_port, "backend_port": backend_port},
                    "tags": tags,
                },
            )
        elif lb_type == "alb":
            write_tfvars(
                Path(args.out) / "30-services" / f"alb-{workload}-{env}-web" / "terraform.tfvars",
                {
                    "resource_group_name": rg,
                    "location": args.location,
                    "vnet_resource_group_name": rg,
                    "vnet_name": vnet,
                    "subnet_name": subnet,
                    "app_gateway_name": name,
                    "frontend_private_ip": frontend_ip,
                    "backend_ip_addresses": [backend],
                    "probe": {"name": "probe-web", "protocol": probe_protocol, "path": "/", "interval": 30, "timeout": 30, "unhealthy_threshold": 3},
                    "listener": {"name": "lstn-web", "protocol": rule_protocol, "frontend_port": frontend_port},
                    "backend_http_settings": {"name": "bhs-web", "protocol": rule_protocol, "port": backend_port, "request_timeout": 30},
                    "sku": {"name": "Standard_v2", "tier": "Standard_v2", "capacity": 1},
                    "tags": tags,
                },
            )


if __name__ == "__main__":
    main()
