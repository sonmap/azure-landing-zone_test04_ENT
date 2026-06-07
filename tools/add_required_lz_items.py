#!/usr/bin/env python3
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def style_sheet(ws, color: str) -> None:
    thin = Side(style="thin", color="D9D9D9")
    fill = PatternFill("solid", fgColor=color)
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(
            64, max(14, len(str(ws.cell(1, col).value or "")) + 6)
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    path = Path("azure_landingzone_terraform_module_design.xlsx")
    wb = openpyxl.load_workbook(path)

    sheets = {
        "21_KMS_KeyVault": (
            "375623",
            [
                "ModuleID", "부서", "업무", "Environment", "ResourceGroup", "KeyVault/HSM Name",
                "Type", "SKU", "CMK 대상", "Private Endpoint", "Public Access", "Purge Protection",
                "Soft Delete Days", "Rotation Days", "Access Model", "tfvars path", "선택 갱신 명령", "Status", "비고",
            ],
            [
                ["MOD-060", "Sales", "영업지원", "dev", "rg-sl-sales-dev", "kv-sl-sales-dev-001", "Key Vault", "standard", "VM Disk/Storage/Secrets", "Y", "Disabled", "Y", 90, 180, "RBAC", "live/30-services/kms-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target kms --workload 영업지원 --environment dev", "Draft", "CMK/secret 저장소, Public 차단"],
                ["MOD-060", "AI Platform", "AI Sandbox", "sandbox", "rg-sl-ai-sandbox", "kv-sl-ai-sbox-001", "Key Vault", "standard", "AI Search/Storage/OpenAI secrets", "Y", "Disabled", "Y", 90, 180, "RBAC", "live/30-services/kms-ai-sandbox/terraform.tfvars", "tools/update_extended_tfvars.sh --target kms --workload \"AI Sandbox\" --environment sandbox", "Draft", "AI 폐쇄망 secret 및 CMK"],
            ],
        ),
        "22_IAM_RBAC_PIM": (
            "1F4E78",
            ["ModuleID", "부서", "업무", "Environment", "Principal Type", "Principal/Group", "Scope Type", "Scope Name", "Role", "PIM", "Eligible/Active", "Duration Hours", "Approver", "MFA/CA 필요", "tfvars path", "선택 갱신 명령", "Status", "비고"],
            [
                ["MOD-061", "Sales", "영업지원", "dev", "Group", "grp-az-sales-dev-operator", "ResourceGroup", "rg-sl-sales-dev", "Reader", "Y", "Eligible", 8, "sales-owner", "Y", "live/00-foundation/iam-rbac/terraform.tfvars", "tools/update_extended_tfvars.sh --target iam --department Sales --environment dev", "Draft", "운영자 조회 권한"],
                ["MOD-061", "Platform", "Network", "prod", "Group", "grp-az-platform-network-admin", "Subscription", "sub-platform", "Network Contributor", "Y", "Eligible", 4, "platform-owner", "Y", "live/00-foundation/iam-rbac/terraform.tfvars", "tools/update_extended_tfvars.sh --target iam --department Platform --environment prod", "Draft", "네트워크 변경 권한"],
            ],
        ),
        "23_ManagedIdentity": (
            "8064A2",
            ["ModuleID", "부서", "업무", "Environment", "Identity Name", "ResourceGroup", "Assigned To", "Federated Credential", "Role 필요", "Scope", "tfvars path", "선택 갱신 명령", "Status", "비고"],
            [
                ["MOD-062", "Sales", "영업지원", "dev", "mi-sl-sales-dev-app", "rg-sl-sales-dev", "Container App/VM", "N", "AcrPull, Key Vault Secrets User", "rg-sl-sales-dev", "live/30-services/identity-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target identity --workload 영업지원 --environment dev", "Draft", "워크로드 관리 ID"],
                ["MOD-062", "Digital", "모바일API", "prod", "mi-sl-digital-prod-aks", "rg-sl-digital-prod-aks", "AKS Workload Identity", "Y", "AcrPull, Key Vault Secrets User", "rg-sl-digital-prod-aks", "live/30-services/identity-digital-prod/terraform.tfvars", "tools/update_extended_tfvars.sh --target identity --workload 모바일API --environment prod", "Draft", "AKS workload identity"],
            ],
        ),
        "24_Backup_DR": (
            "9E480E",
            ["ModuleID", "부서", "업무", "Environment", "대상", "Vault Name", "ResourceGroup", "Policy", "Retention Daily", "Retention Monthly", "DR Region", "RPO", "RTO", "tfvars path", "선택 갱신 명령", "Status", "비고"],
            [
                ["MOD-063", "Sales", "영업지원", "dev", "VM", "rsv-sl-sales-dev-001", "rg-sl-sales-dev", "Daily-30D", 30, 12, "koreasouth", "24h", "4h", "live/30-services/backup-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target backup --workload 영업지원 --environment dev", "Draft", "VM 백업"],
                ["MOD-063", "AI Platform", "AI Sandbox", "sandbox", "Storage", "bv-sl-ai-sbox-001", "rg-sl-ai-sandbox", "Blob-Operational", 30, 6, "koreasouth", "24h", "8h", "live/30-services/backup-ai-sandbox/terraform.tfvars", "tools/update_extended_tfvars.sh --target backup --workload \"AI Sandbox\" --environment sandbox", "Draft", "AI 데이터 보호"],
            ],
        ),
        "25_Monitoring_Alert": (
            "C00000",
            ["ModuleID", "부서", "업무", "Environment", "Workspace Name", "ResourceGroup", "Diagnostic 대상", "Log Category", "Metric Alert", "Action Group", "Sentinel", "Retention Days", "tfvars path", "선택 갱신 명령", "Status", "비고"],
            [
                ["MOD-064", "Platform", "공통", "prod", "law-sl-platform-prod", "rg-sl-monitoring-krc", "All RG/Firewall/AKS/KeyVault", "Audit/AllMetrics", "CPU,Availability,DeniedTraffic", "ag-sl-soc", "Y", 180, "live/10-platform/monitoring/terraform.tfvars", "tools/update_extended_tfvars.sh --target monitoring --department Platform --environment prod", "Draft", "SOC/Sentinel 연동"],
                ["MOD-064", "Sales", "영업지원", "dev", "law-sl-sales-dev", "rg-sl-sales-dev", "VM/LB/AppGW", "AllLogs/AllMetrics", "VM down, LB probe fail", "ag-sl-sales-dev", "N", 90, "live/30-services/monitoring-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target monitoring --workload 영업지원 --environment dev", "Draft", "업무 모니터링"],
            ],
        ),
        "26_WAF_DDoS": (
            "7030A0",
            ["ModuleID", "부서", "업무", "Environment", "Control Type", "Name", "ResourceGroup", "Mode/SKU", "연결 대상", "Policy Rule Set", "Public 허용", "tfvars path", "선택 갱신 명령", "Status", "비고"],
            [
                ["MOD-065", "Sales", "영업지원", "dev", "WAF Policy", "waf-sl-sales-dev-web", "rg-sl-sales-dev", "Prevention", "agw-sl-sales-dev-web", "OWASP 3.2", "N", "live/30-services/waf-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target waf --workload 영업지원 --environment dev", "Draft", "Internal AppGW WAF"],
                ["MOD-065", "Platform", "Network", "prod", "DDoS Plan", "ddos-sl-platform-krc", "rg-sl-hub-krc", "Network Protection", "vnet-sl-hub-krc", "N/A", "N", "live/10-platform/ddos/terraform.tfvars", "tools/update_extended_tfvars.sh --target ddos --department Platform --environment prod", "Draft", "필요 시만 활성화, 비용 주의"],
            ],
        ),
        "27_PrivateEndpoint_Request": (
            "31859B",
            ["ModuleID", "부서", "업무", "Environment", "대상 서비스", "Resource Name", "ResourceGroup", "VNet", "Subnet", "Private DNS Zone", "Manual Approval", "Public Access", "tfvars path", "선택 갱신 명령", "Status", "비고"],
            [
                ["MOD-066", "Sales", "영업지원", "dev", "Storage Blob", "stslssalesdev001", "rg-sl-sales-dev", "vnet-sl-sales-dev", "snet-pe", "privatelink.blob.core.windows.net", "N", "Disabled", "live/40-access/private-endpoint-sales-dev-storage/terraform.tfvars", "tools/update_extended_tfvars.sh --target private-endpoint --workload 영업지원 --environment dev", "Draft", "스토리지 사설 접근"],
                ["MOD-066", "AI Platform", "AI Sandbox", "sandbox", "Azure OpenAI", "oai-sl-ai-sbox-001", "rg-sl-ai-sandbox", "vnet-sl-ai-sandbox", "snet-pe", "privatelink.openai.azure.com", "N", "Disabled", "live/40-access/private-endpoint-ai-sandbox-openai/terraform.tfvars", "tools/update_extended_tfvars.sh --target private-endpoint --workload \"AI Sandbox\" --environment sandbox", "Draft", "OpenAI 사설 접근"],
            ],
        ),
        "28_Compliance_Policy": (
            "595959",
            ["ModuleID", "정책명", "Scope", "Assignment Name", "Effect", "대상", "예외", "Initiative", "tfvars path", "선택 갱신 명령", "Status", "비고"],
            [
                ["MOD-067", "Deny Public IP", "ManagementGroup", "deny-public-ip", "Deny", "Public IP 생성", "Firewall 검증 예외 필요", "SamsungLife-Secure-LZ", "live/00-foundation/policy/terraform.tfvars", "tools/update_extended_tfvars.sh --target policy", "Draft", "인터넷 직접 노출 차단"],
                ["MOD-067", "Require Private Endpoint", "ManagementGroup", "audit-private-endpoint", "Audit", "PaaS 서비스", "없음", "SamsungLife-Secure-LZ", "live/00-foundation/policy/terraform.tfvars", "tools/update_extended_tfvars.sh --target policy", "Draft", "PaaS 사설 접근 강제"],
            ],
        ),
        "29_Operations_Runbook": (
            "4F81BD",
            ["RunbookID", "운영 시나리오", "대상", "명령/절차", "승인자", "Rollback", "점검 항목", "비고"],
            [
                ["RUN-001", "Excel 변경 후 선택 반영", "tfvars", "tools/update_selected_tfvars.sh 또는 tools/update_extended_tfvars.sh 실행", "서비스오너", "백업 xlsx/tfvars 복원", "terraform plan 확인", "apply 전 CAB 승인"],
                ["RUN-002", "1시간 검증 후 삭제", "Terraform root", "terraform destroy 또는 예약 destroy 스크립트", "플랫폼오너", "재 apply", "az resource list 확인", "고비용 리소스 우선 삭제"],
            ],
        ),
        "30_ContainerApps": (
            "00B050",
            ["ModuleID", "부서", "업무", "Environment", "Container App Name", "ResourceGroup", "VNet", "Subnet", "Image", "CPU", "Memory", "Min Replicas", "Max Replicas", "Ingress", "Target Port", "Managed Identity", "ACR", "Secrets", "Env Vars", "Dapr", "Log Analytics", "Public Access", "tfvars path", "선택 갱신 명령", "Status", "비고"],
            [
                ["MOD-070", "Sales", "영업지원", "dev", "ca-sl-sales-dev-api", "rg-sl-sales-dev", "vnet-sl-sales-dev", "snet-app", "acrslplatform.azurecr.io/sales-api:dev", 0.5, "1Gi", 0, 2, "internal", 8080, "mi-sl-sales-dev-app", "acrslplatform", "kv-ref:db-password", "ENV=dev", "N", "law-sl-sales-dev", "Disabled", "live/30-services/containerapp-sales-dev/terraform.tfvars", "tools/update_extended_tfvars.sh --target container-app --workload 영업지원 --environment dev", "Draft", "GCP Cloud Run 대응"],
                ["MOD-070", "AI Platform", "AI Sandbox", "sandbox", "ca-sl-ai-sbox-rag", "rg-sl-ai-sandbox", "vnet-sl-ai-sandbox", "snet-app", "acrslplatform.azurecr.io/rag-api:sbox", 1.0, "2Gi", 0, 2, "internal", 8080, "mi-sl-ai-sbox-app", "acrslplatform", "kv-ref:openai-key", "ENV=sandbox", "N", "law-sl-platform-prod", "Disabled", "live/30-services/containerapp-ai-sandbox/terraform.tfvars", "tools/update_extended_tfvars.sh --target container-app --workload \"AI Sandbox\" --environment sandbox", "Draft", "Private RAG API"],
            ],
        ),
    }

    for name, (color, headers, rows) in sheets.items():
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        style_sheet(ws, color)

    if "19_TFVars_Mapping" in wb.sheetnames:
        mapping = wb["19_TFVars_Mapping"]
        existing = {str(mapping.cell(r, 4).value) for r in range(2, mapping.max_row + 1)}
        number = mapping.max_row
        for sheet_name in [k for k in sheets if k != "29_Operations_Runbook"]:
            ws = wb[sheet_name]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            for row_idx in range(2, ws.max_row + 1):
                row = {headers[c - 1]: ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)}
                tfvars_path = row.get("tfvars path")
                command = row.get("선택 갱신 명령")
                if not tfvars_path or str(tfvars_path) in existing:
                    continue
                condition = []
                for key in ["부서", "업무", "Environment", "Name", "Container App Name", "Control Type", "대상 서비스", "정책명"]:
                    if row.get(key):
                        condition.append(f"{key}={row[key]}")
                number += 1
                mapping.append([
                    number,
                    sheet_name,
                    ", ".join(condition),
                    tfvars_path,
                    command or "",
                    str(tfvars_path).rsplit("/terraform.tfvars", 1)[0],
                    row.get("비고", ""),
                ])

    wb.save(path)
    print("extended landing zone sheets added")


if __name__ == "__main__":
    main()
