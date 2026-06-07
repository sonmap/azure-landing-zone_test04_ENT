#!/usr/bin/env python3
"""
Azure Landing Zone Excel 설계서 -> Terraform .tfvars 생성기.

기본 사용:
  python tools/excel_to_tfvars.py ^
    --excel azure_landingzone_terraform_module_design.xlsx ^
    --out generated_tfvars ^
    --tenant-id <tenant_id> ^
    --subscription-id <subscription_id>

기본 출력은 HCL 형식의 terraform.tfvars 입니다. JSON이 필요하면 --format json을 사용합니다.
Excel 셀에는 ${subscription_id}, ${tenant_id}, ${location}, ${hub_vnet_id} 같은 값을 넣어
CLI 변수로 치환할 수 있습니다.
"""
from __future__ import annotations

import argparse
import json
import re
from ipaddress import ip_network
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("openpyxl이 필요합니다. pip install openpyxl 후 실행하십시오.") from exc


KNOWN_SLUGS = {
    "영업": "sales",
    "영업지원": "sales",
    "계약": "policy",
    "계약관리": "policy",
    "상품": "product",
    "디지털": "digital",
    "모바일api": "digital",
    "ai platform": "ai",
    "ai sandbox": "ai",
    "관리": "mgmt",
    "공통": "shared",
}


def rows_as_dicts(ws) -> List[Dict[str, Any]]:
    headers: Optional[List[str]] = None
    result: List[Dict[str, Any]] = []
    header_keys = {
        "moduleid",
        "id",
        "networkid",
        "requestid",
        "request_id",
        "항목",
        "모듈id",
        "구분",
    }

    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if not any(v is not None for v in values):
            continue
        if headers is None:
            normalized = {str(v).strip().lower().replace(" ", "") for v in values if v is not None}
            if normalized.intersection(header_keys):
                headers = [str(v).strip() if v is not None else f"col_{i}" for i, v in enumerate(values)]
            continue

        item = {}
        for i in range(min(len(headers), len(values))):
            if headers[i] and values[i] is not None:
                item[headers[i]] = values[i]
        if item:
            result.append(item)
    return result


def value(row: Dict[str, Any], *keys: str, default: Any = None) -> Any:
    lowered = {str(k).strip().lower().replace(" ", "").replace("_", ""): k for k in row.keys()}
    for key in keys:
        normalized = key.strip().lower().replace(" ", "").replace("_", "")
        real_key = lowered.get(normalized)
        if real_key is not None and row[real_key] not in (None, ""):
            return row[real_key]
    return default


def as_bool(raw: Any, default: bool = False) -> bool:
    if raw is None or raw == "":
        return default
    text = str(raw).strip().lower()
    return text in {"y", "yes", "true", "1", "enabled", "enable", "사용", "예", "필요"}


def as_int(raw: Any, default: int) -> int:
    try:
        return int(float(str(raw).strip()))
    except (TypeError, ValueError):
        return default


def slugify(raw: Any, fallback: str = "item") -> str:
    text = str(raw or "").strip()
    mapped = KNOWN_SLUGS.get(text.lower()) or KNOWN_SLUGS.get(text)
    if mapped:
        return mapped
    text = text.lower()
    text = re.sub(r"[^a-z0-9가-힣]+", "-", text).strip("-")
    parts = [KNOWN_SLUGS.get(p, p) for p in text.split("-") if p]
    ascii_text = re.sub(r"[^a-z0-9-]+", "", "-".join(parts)).strip("-")
    return ascii_text or fallback


def subnet_name(raw: Any) -> str:
    text = str(raw or "").strip()
    if not text:
        return "snet-app"
    if text.lower().startswith("snet-"):
        return text
    return f"snet-{slugify(text, 'app')}"


def render_vars(obj: Any, context: Dict[str, str]) -> Any:
    if isinstance(obj, str):
        def replace(match: re.Match[str]) -> str:
            return str(context.get(match.group(1), match.group(0)))

        return re.sub(r"\$\{([A-Za-z0-9_]+)\}", replace, obj)
    if isinstance(obj, list):
        return [render_vars(v, context) for v in obj]
    if isinstance(obj, dict):
        return {k: render_vars(v, context) for k, v in obj.items()}
    return obj


def hcl_string(text: str) -> str:
    return json.dumps(str(text), ensure_ascii=False)


def hcl_key(key: Any) -> str:
    text = str(key)
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", text):
        return text
    return hcl_string(text)


def to_hcl(obj: Any, indent: int = 0) -> str:
    pad = " " * indent
    child = " " * (indent + 2)
    if isinstance(obj, bool):
        return "true" if obj else "false"
    if isinstance(obj, (int, float)):
        return str(obj)
    if obj is None:
        return "null"
    if isinstance(obj, str):
        return hcl_string(obj)
    if isinstance(obj, list):
        if not obj:
            return "[]"
        return "[\n" + ",\n".join(f"{child}{to_hcl(v, indent + 2)}" for v in obj) + f"\n{pad}]"
    if isinstance(obj, dict):
        if not obj:
            return "{}"
        lines = ["{"]
        for key in sorted(obj.keys()):
            lines.append(f"{child}{hcl_key(key)} = {to_hcl(obj[key], indent + 2)}")
        lines.append(f"{pad}}}")
        return "\n".join(lines)
    return hcl_string(str(obj))



def parse_note_params(raw: Any) -> Dict[str, str]:
    params: Dict[str, str] = {}
    for part in str(raw or "").split(";"):
        if "=" not in part:
            continue
        key, val = part.split("=", 1)
        key = key.strip()
        if key:
            params[key] = val.strip()
    return params


def write_tfvars(path: Path, data: Dict[str, Any], context: Dict[str, str], fmt: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rendered = render_vars(data, context)
    if fmt == "json":
        path = path.with_name("terraform.auto.tfvars.json")
        path.write_text(json.dumps(rendered, indent=2, ensure_ascii=False), encoding="utf-8")
    else:
        lines = [f"{key} = {to_hcl(rendered[key])}" for key in sorted(rendered.keys())]
        path.write_text("\n\n".join(lines) + "\n", encoding="utf-8")
    print(f"created: {path}")


def parse_tags(row: Dict[str, Any], base: Dict[str, str]) -> Dict[str, str]:
    tags = dict(base)
    aliases = {
        "department": ("부서", "department", "dept"),
        "env": ("Environment", "env"),
        "owner": ("Tag:owner", "owner"),
        "costcenter": ("Tag:costcenter", "costcenter", "CostCenter"),
        "data_class": ("Data Class", "데이터등급", "data_class"),
        "itsm_ticket": ("RequestID", "ITSM", "itsm_ticket"),
    }
    for tag, keys in aliases.items():
        raw = value(row, *keys)
        if raw not in (None, ""):
            tags[tag] = str(raw)
    tags.setdefault("managed_by", "terraform")
    return tags


def id_for_vnet(subscription_id: str, rg: str, vnet: str) -> str:
    return f"/subscriptions/{subscription_id}/resourceGroups/{rg}/providers/Microsoft.Network/virtualNetworks/{vnet}"


def id_for_subnet(subscription_id: str, rg: str, vnet: str, subnet: str) -> str:
    return f"{id_for_vnet(subscription_id, rg, vnet)}/subnets/{subnet}"


def derived_subnets(address_space: str, names: Iterable[str]) -> Dict[str, Dict[str, Any]]:
    try:
        network = ip_network(address_space, strict=False)
        blocks = list(network.subnets(new_prefix=min(network.prefixlen + 4, 28)))
    except ValueError:
        blocks = []

    result: Dict[str, Dict[str, Any]] = {}
    for idx, raw_name in enumerate(names):
        name = subnet_name(raw_name)
        prefix = str(blocks[idx]) if idx < len(blocks) else "10.0.0.0/24"
        result[slugify(name, f"snet{idx + 1}")] = {
            "name": name,
            "address_prefixes": [prefix],
            "create_nsg": True,
            "associate_route_table": name.lower() not in {"snet-pe", "snet-private-endpoint"},
            "private_endpoint_network_policies": "Disabled" if "pe" in name.lower() else "Enabled",
        }
    return result


def workload_index(workloads: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    idx: Dict[str, Dict[str, Any]] = {}
    for row in workloads:
        env = str(value(row, "Environment", default="dev")).lower()
        business = slugify(value(row, "업무", default=value(row, "부서", default="workload")))
        dept = slugify(value(row, "부서", default=business))
        for key in {f"{business}:{env}", f"{dept}:{env}"}:
            idx[key] = row
    return idx


def find_workload(index: Dict[str, Dict[str, Any]], row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    env = str(value(row, "Environment", default="dev")).lower()
    business = slugify(value(row, "업무", "AI 업무", default=value(row, "부서", default="workload")))
    dept = slugify(value(row, "부서", default=business))
    return index.get(f"{business}:{env}") or index.get(f"{dept}:{env}")


def vm_image(os_name: Any) -> Dict[str, str]:
    text = str(os_name or "").lower()
    if "win" in text:
        return {"image_publisher": "MicrosoftWindowsServer", "image_offer": "WindowsServer", "image_sku": "2022-datacenter-g2"}
    if "ubuntu" in text:
        return {"image_publisher": "Canonical", "image_offer": "0001-com-ubuntu-server-jammy", "image_sku": "22_04-lts-gen2"}
    return {"image_publisher": "RedHat", "image_offer": "RHEL", "image_sku": "9-lvm-gen2"}


def vm_bucket(role: Any) -> str:
    text = slugify(role, "agent")
    if "web" in text:
        return "web_vms"
    if "was" in text or "app" in text:
        return "was_vms"
    if "db" in text:
        return "db_vms"
    return "agent_vms"


def build_common_args(args: argparse.Namespace) -> Dict[str, Any]:
    return {
        "tenant_id": args.tenant_id,
        "subscription_id": args.subscription_id,
        "location": args.location,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--out", default="generated_tfvars")
    parser.add_argument("--format", choices=["hcl", "json"], default="hcl")
    parser.add_argument("--tenant-id", required=True)
    parser.add_argument("--subscription-id", required=True)
    parser.add_argument("--location", default="koreacentral")
    parser.add_argument("--hub-resource-group-name", default="rg-sl-hub-krc")
    parser.add_argument("--hub-vnet-name", default="vnet-sl-hub-krc")
    parser.add_argument("--hub-vnet-cidr", default="10.39.0.0/20")
    parser.add_argument("--hub-dns-inbound-ip", default="10.39.1.4")
    parser.add_argument("--firewall-private-ip", default="10.39.0.68")
    parser.add_argument("--ssh-public-key", default="ssh-rsa REPLACE_ME")
    parser.add_argument("--aks-log-analytics-workspace-id", default="/subscriptions/${subscription_id}/resourceGroups/rg-monitoring/providers/Microsoft.OperationalInsights/workspaces/law-replace")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    out = Path(args.out)
    common = build_common_args(args)
    context = {
        "tenant_id": args.tenant_id,
        "subscription_id": args.subscription_id,
        "location": args.location,
        "hub_resource_group_name": args.hub_resource_group_name,
        "hub_vnet_name": args.hub_vnet_name,
        "hub_vnet_id": id_for_vnet(args.subscription_id, args.hub_resource_group_name, args.hub_vnet_name),
        "hub_dns_inbound_ip": args.hub_dns_inbound_ip,
        "firewall_private_ip": args.firewall_private_ip,
    }
    base_tags = {"project": "samsunglife-landingzone", "managed_by": "terraform"}

    workloads = rows_as_dicts(wb["04_Workload_LZ"]) if "04_Workload_LZ" in wb.sheetnames else []
    workload_by_key = workload_index(workloads)
    if workloads:
        resource_groups: Dict[str, Dict[str, Any]] = {
            "hub": {"name": args.hub_resource_group_name, "location": args.location, "tags": base_tags}
        }
        for row in workloads:
            rg = str(value(row, "ResourceGroup", "resource_group_name", default="rg-replace"))
            key = slugify(rg, f"rg{len(resource_groups)}")
            resource_groups[key] = {"name": rg, "location": args.location, "tags": parse_tags(row, base_tags)}
        write_tfvars(
            out / "00-foundation" / "resource-groups" / "terraform.tfvars",
            {**common, "common_tags": base_tags, "resource_groups": resource_groups},
            context,
            args.format,
        )

    write_tfvars(
        out / "10-platform" / "hub-network" / "terraform.tfvars",
        {
            **common,
            "common_tags": base_tags,
            "resource_group_name": args.hub_resource_group_name,
            "hub_name": args.hub_vnet_name,
            "address_space": [args.hub_vnet_cidr],
            "dns_servers": [],
            "enable_private_dns_resolver": True,
            "dns_inbound_subnet_key": "dns_in",
            "dns_outbound_subnet_key": "dns_out",
            "subnets": {
                "gateway": {"name": "GatewaySubnet", "address_prefixes": ["10.39.0.0/27"], "create_nsg": False},
                "firewall": {"name": "AzureFirewallSubnet", "address_prefixes": ["10.39.0.64/26"], "create_nsg": False},
                "dns_in": {"name": "snet-dns-inbound", "address_prefixes": ["10.39.1.0/28"], "create_nsg": False, "delegate_dns_resolver": True},
                "dns_out": {"name": "snet-dns-outbound", "address_prefixes": ["10.39.1.16/28"], "create_nsg": False, "delegate_dns_resolver": True},
                "bastion": {"name": "AzureBastionSubnet", "address_prefixes": ["10.39.2.0/26"], "create_nsg": False},
                "pe": {"name": "snet-shared-pe", "address_prefixes": ["10.39.3.0/24"], "create_nsg": True},
            },
        },
        context,
        args.format,
    )

    network_rows = rows_as_dicts(wb["03_Network_Connectivity"]) if "03_Network_Connectivity" in wb.sheetnames else []
    for row in network_rows:
        if str(value(row, "구분", default="")).strip().lower() != "azure firewall":
            continue
        params = parse_note_params(value(row, "비고", default=""))
        firewall_name = str(value(row, "이름", default="afw-sl-hub-krc-001"))
        write_tfvars(
            out / "10-platform" / "azure-firewall" / "terraform.tfvars",
            {
                **common,
                "resource_group_name": params.get("resource_group_name", args.hub_resource_group_name),
                "firewall_name": firewall_name,
                "firewall_subnet_id": str(value(row, "CIDR/대상", default=id_for_subnet(args.subscription_id, args.hub_resource_group_name, args.hub_vnet_name, "AzureFirewallSubnet"))),
                "public_ip_name": params.get("public_ip_name", f"pip-{firewall_name}"),
                "sku_tier": params.get("sku_tier", "Standard"),
                "common_tags": base_tags,
            },
            context,
            args.format,
        )
        break

    if workloads:
        for row in workloads:
            env = str(value(row, "Environment", default="dev")).lower()
            dept = slugify(value(row, "부서", default="dept"))
            rg = str(value(row, "ResourceGroup", default="rg-replace"))
            vnet = str(value(row, "VNet", default=f"vnet-sl-{dept}-{env}"))
            cidr = str(value(row, "CIDR", default="10.40.0.0/20"))
            subnet_list = str(value(row, "Subnet 구성", default="web/was/db/pe")).split("/")
            write_tfvars(
                out / "20-workload" / f"{dept}-{env}-spoke" / "terraform.tfvars",
                {
                    **common,
                    "common_tags": parse_tags(row, base_tags),
                    "resource_group_name": rg,
                    "spoke_name": vnet,
                    "address_space": [cidr],
                    "dns_servers": [args.hub_dns_inbound_ip],
                    "hub_vnet_id": "${hub_vnet_id}",
                    "hub_resource_group_name": args.hub_resource_group_name,
                    "hub_vnet_name": args.hub_vnet_name,
                    "firewall_private_ip": args.firewall_private_ip,
                    "subnets": derived_subnets(cidr, subnet_list),
                },
                context,
                args.format,
            )

    if "05_VM_Module" in wb.sheetnames:
        groups: Dict[str, Dict[str, Any]] = {}
        for row in rows_as_dicts(wb["05_VM_Module"]):
            vm_name = str(value(row, "VM Name", "VM명", "name", default="")).strip()
            if not vm_name:
                continue
            workload = find_workload(workload_by_key, row)
            env = str(value(row, "Environment", default="dev")).lower()
            business = slugify(value(row, "업무", default="workload"))
            key = f"{business}-{env}"
            group = groups.setdefault(
                key,
                {
                    **common,
                    "resource_group_name": str(value(workload or {}, "ResourceGroup", default=f"rg-sl-{business}-{env}")),
                    "location": args.location,
                    "vnet_resource_group_name": str(value(workload or {}, "ResourceGroup", default=f"rg-sl-{business}-{env}")),
                    "vnet_name": str(value(workload or {}, "VNet", default=f"vnet-sl-{business}-{env}")),
                    "admin_username": "azureuser",
                    "ssh_public_key": args.ssh_public_key,
                    "os_disk_storage_type": "Premium_LRS",
                    "common_tags": parse_tags(row, base_tags),
                    "web_vms": {},
                    "was_vms": {},
                    "db_vms": {},
                    "agent_vms": {},
                    **vm_image(value(row, "OS", default="RHEL")),
                    "image_version": "latest",
                },
            )
            bucket = vm_bucket(value(row, "역할", "role", default="agent"))
            group[bucket][slugify(vm_name, "vm")] = {
                "name": vm_name,
                "subnet_name": subnet_name(value(row, "Subnet", default="app")),
                "private_ip_address": str(value(row, "Private IP", default="")),
                "vm_size": str(value(row, "Size", default="Standard_D2s_v3")),
                "os_disk_size_gb": as_int(value(row, "Disk GB", default=128), 128),
                "role": str(value(row, "역할", default=bucket.replace("_vms", ""))),
                "itsm_ticket": str(value(row, "RequestID", default="CSR-REPLACE")),
                "data_disks": {},
            }
        for key, data in groups.items():
            write_tfvars(out / "30-services" / f"vm-{key}" / "terraform.tfvars", data, context, args.format)

    if "07_AKS_Module" in wb.sheetnames:
        for row in rows_as_dicts(wb["07_AKS_Module"]):
            cluster_name = str(value(row, "AKS Cluster", default="")).strip()
            if not cluster_name:
                continue
            workload = find_workload(workload_by_key, row)
            env = str(value(row, "Environment", default="dev")).lower()
            business = slugify(value(row, "업무", default="aks"))
            write_tfvars(
                out / "30-services" / f"aks-{business}-{env}" / "terraform.tfvars",
                {
                    "resource_group_name": str(value(workload or {}, "ResourceGroup", default=f"rg-sl-{business}-{env}")),
                    "location": args.location,
                    "tags": parse_tags(row, base_tags),
                    "vnet_resource_group_name": str(value(workload or {}, "ResourceGroup", default=f"rg-sl-{business}-{env}")),
                    "vnet_name": str(value(workload or {}, "VNet", default=f"vnet-sl-{business}-{env}")),
                    "aks_subnet_name": "snet-aks",
                    "cluster": {
                        "name": cluster_name,
                        "kubernetes_version": "1.30",
                        "dns_prefix": slugify(cluster_name, "aks"),
                        "private_cluster_enabled": True,
                        "sku_tier": "Free",
                        "local_account_disabled": True,
                    },
                    "default_node_pool": {
                        "name": "system",
                        "vm_size": str(value(row, "NodePool", default="Standard_D2s_v3")),
                        "node_count": 1,
                        "auto_scaling_enabled": True,
                        "min_count": 1,
                        "max_count": 2,
                        "os_disk_size_gb": 64,
                    },
                    "user_node_pools": {},
                    "network_profile": {
                        "network_plugin": "azure",
                        "network_policy": "azure",
                        "service_cidr": "10.250.0.0/16",
                        "dns_service_ip": "10.250.0.10",
                        "outbound_type": "userDefinedRouting",
                    },
                    "monitoring": {"enabled": True, "log_analytics_workspace_id": args.aks_log_analytics_workspace_id},
                },
                context,
                args.format,
            )

    if "08_AI_Private_Module" in wb.sheetnames:
        for row in rows_as_dicts(wb["08_AI_Private_Module"]):
            dept = slugify(value(row, "부서", default="ai"))
            env = str(value(row, "Environment", default="sandbox")).lower()
            workload = find_workload(workload_by_key, row)
            rg = str(value(workload or {}, "ResourceGroup", default=f"rg-sl-{dept}-ai-{env}"))
            vnet = str(value(workload or {}, "VNet", default=f"vnet-sl-{dept}-{env}"))
            storage_name = re.sub(r"[^a-z0-9]", "", f"stsl{dept}{env}001")[:24]
            write_tfvars(
                out / "30-services" / f"ai-{dept}-{env}" / "terraform.tfvars",
                {
                    "resource_group_name": rg,
                    "location": args.location,
                    "tags": parse_tags(row, base_tags),
                    "vnet_resource_group_name": rg,
                    "vnet_name": vnet,
                    "private_endpoint_subnet_name": "snet-pe",
                    "private_dns_zone_resource_group_name": args.hub_resource_group_name,
                    "openai": {
                        "enabled": as_bool(value(row, "OpenAI", default=True), True),
                        "name": f"oai-sl-{dept}-{env}-001",
                        "sku_name": "S0",
                        "public_network_access": "Disabled",
                        "custom_subdomain_name": f"oai-sl-{dept}-{env}-001",
                    },
                    "search": {
                        "enabled": as_bool(value(row, "AI Search", default=True), True),
                        "name": f"srch-sl-{dept}-{env}-001",
                        "sku": "basic",
                        "replica_count": 1,
                        "partition_count": 1,
                        "public_network_access": "Disabled",
                    },
                    "storage": {
                        "enabled": as_bool(value(row, "Storage", default=True), True),
                        "name": storage_name,
                        "account_tier": "Standard",
                        "account_replication_type": "LRS",
                        "public_network_access": "Disabled",
                        "allow_blob_public_access": False,
                        "default_action": "Deny",
                        "containers": ["landing", "rag", "audit"],
                    },
                    "keyvault": {
                        "enabled": as_bool(value(row, "KeyVault", default=True), True),
                        "name": f"kv-sl-{dept}-{env}-001"[:24],
                        "sku_name": "standard",
                        "tenant_id": args.tenant_id,
                        "public_network_access": "Disabled",
                        "purge_protection": True,
                    },
                    "private_endpoints": {"openai": True, "search": True, "blob": True, "keyvault": True},
                    "private_dns_zones": {
                        "openai": "privatelink.openai.azure.com",
                        "search": "privatelink.search.windows.net",
                        "blob": "privatelink.blob.core.windows.net",
                        "keyvault": "privatelink.vaultcore.azure.net",
                    },
                },
                context,
                args.format,
            )

    write_tfvars(
        out / "40-access" / "private-dns-zones" / "terraform.tfvars",
        {
            **common,
            "resource_group_name": args.hub_resource_group_name,
            "zones": [
                "privatelink.openai.azure.com",
                "privatelink.search.windows.net",
                "privatelink.blob.core.windows.net",
                "privatelink.vaultcore.azure.net",
            ],
            "virtual_network_links": {
                "hub-openai": {
                    "zone_name": "privatelink.openai.azure.com",
                    "name": "lnk-hub-openai",
                    "virtual_network_id": "${hub_vnet_id}",
                    "registration_enabled": False,
                },
                "hub-search": {
                    "zone_name": "privatelink.search.windows.net",
                    "name": "lnk-hub-search",
                    "virtual_network_id": "${hub_vnet_id}",
                    "registration_enabled": False,
                },
                "hub-blob": {
                    "zone_name": "privatelink.blob.core.windows.net",
                    "name": "lnk-hub-blob",
                    "virtual_network_id": "${hub_vnet_id}",
                    "registration_enabled": False,
                },
                "hub-keyvault": {
                    "zone_name": "privatelink.vaultcore.azure.net",
                    "name": "lnk-hub-keyvault",
                    "virtual_network_id": "${hub_vnet_id}",
                    "registration_enabled": False,
                },
            },
            "common_tags": base_tags,
        },
        context,
        args.format,
    )

    print("완료. 생성된 terraform.tfvars를 검토한 뒤 필요한 live root로 복사하고 terraform plan을 수행하십시오.")


if __name__ == "__main__":
    main()
